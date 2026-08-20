from __future__ import annotations

from contextlib import asynccontextmanager
from datetime import datetime
from io import BytesIO
import asyncio
import math
import os
import logging
from pathlib import Path
import shutil
import time
from urllib.parse import quote

from fastapi import Depends, FastAPI, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import func, or_, select, text
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.orm import Session, selectinload

from .config import get_settings
from .database import Base, SessionLocal, engine, get_db
from .models import (
    AppSetting,
    MappingStrategy,
    MatchStatus,
    SelectionCandidate,
    SelectionCandidateFile,
    SelectionStatus,
    utcnow,
)
from .schemas import DecisionRequest, ScanRequest, SettingsPayload
from .services import (
    candidate_query,
    configured_data_roots,
    CopyConflictError,
    discover_files,
    file_url,
    finalize_copies,
    get_or_create_settings,
    copy_candidate_files,
    read_label_json,
    resolve_directory,
    resolve_safe,
    rollback_copies,
    scan_candidates,
    selection_summary,
    find_thumbnail,
    queue_thumbnail_generation,
    thumbnail_state,
    thumbnail_job_status,
)

settings = get_settings()
logger = logging.getLogger(__name__)


@asynccontextmanager
async def lifespan(_: FastAPI):
    for path in (
        settings.data_root_path,
        settings.thumbnail_root_path,
        settings.selected_data_path,
        settings.app_data_path,
    ):
        path.mkdir(parents=True, exist_ok=True)
    Base.metadata.create_all(bind=engine)
    # 기존 SQLite 파일도 그대로 사용할 수 있도록 새 설정 컬럼을 보강한다.
    with engine.begin() as connection:
        columns = {row[1] for row in connection.execute(text("PRAGMA table_info(app_settings)"))}
        if "raw_relative_path" not in columns:
            connection.execute(text("ALTER TABLE app_settings ADD COLUMN raw_relative_path TEXT NOT NULL DEFAULT ''"))
        if "labeled_relative_path" not in columns:
            connection.execute(text("ALTER TABLE app_settings ADD COLUMN labeled_relative_path TEXT NOT NULL DEFAULT ''"))
        if "thumbnail_relative_path" not in columns:
            connection.execute(text("ALTER TABLE app_settings ADD COLUMN thumbnail_relative_path TEXT NOT NULL DEFAULT ''"))
        if "paths_configured" not in columns:
            connection.execute(text("ALTER TABLE app_settings ADD COLUMN paths_configured INTEGER NOT NULL DEFAULT 0"))
        if "annotation_method_code" not in columns:
            connection.execute(text("ALTER TABLE app_settings ADD COLUMN annotation_method_code TEXT NOT NULL DEFAULT 'bbox_2d'"))
    with SessionLocal() as db:
        get_or_create_settings(db)
    if settings.thumbnail_error:
        logger.error(settings.thumbnail_error)
    yield


app = FastAPI(title="Upcycling Prework Selector", version="1.0.0", lifespan=lifespan)
app.add_middleware(
    CORSMiddleware,
    allow_origins=settings.cors_origin_list,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


def _serialize_file(item: SelectionCandidateFile, *, thumbnail_relative_path: str = "", raw_root: Path | None = None) -> dict:
    payload = {
        "id": item.id,
        "file_group": item.file_group,
        "reference_order": item.reference_order,
        "original_relative_path": item.original_relative_path,
        "selected_relative_path": item.selected_relative_path,
        "extension": item.extension,
        "size": item.size,
        "mtime": item.mtime,
        "is_previewable_image": item.is_previewable_image,
        "file_url": file_url(item),
    }
    if item.file_group == "raw" and raw_root is not None:
        raw_path = resolve_safe(raw_root, item.original_relative_path, must_exist=False)
        state, found = thumbnail_state(settings, thumbnail_relative_path, raw_root, item.original_relative_path, raw_path)
        payload["thumbnail_url"] = f"/api/thumbnails?path={quote(item.original_relative_path, safe='')}" if found else None
        payload["thumbnail_status"] = state
    return payload


def _serialize_candidate(candidate: SelectionCandidate, *, include_label: bool = False, labeled_root: Path | None = None, raw_root: Path | None = None, thumbnail_relative_path: str = "") -> dict:
    payload = {
        "id": candidate.id,
        "fingerprint": candidate.fingerprint,
        "match_key": candidate.match_key,
        "mapping_strategy": candidate.mapping_strategy,
        "match_status": candidate.match_status,
        "selection_status": candidate.selection_status,
        "error_message": candidate.error_message,
        "created_at": candidate.created_at,
        "updated_at": candidate.updated_at,
        "decided_at": candidate.decided_at,
        "files": [_serialize_file(item, raw_root=raw_root, thumbnail_relative_path=thumbnail_relative_path) for item in candidate.files],
    }
    payload["label_json"] = read_label_json(candidate, settings, labeled_root) if include_label and labeled_root else None
    return payload


def _candidate_or_404(db: Session, candidate_id: str) -> SelectionCandidate:
    candidate = db.scalar(
        candidate_query().where(SelectionCandidate.id == candidate_id)
    )
    if candidate is None:
        raise HTTPException(status_code=404, detail="선별 후보를 찾을 수 없습니다.")
    return candidate


@app.get("/health")
def health() -> dict:
    return {"status": "ok", "version": app.version}


@app.get("/api/settings")
def retrieve_settings(db: Session = Depends(get_db)) -> dict:
    row = get_or_create_settings(db)
    return {
        "mapping_strategy": row.mapping_strategy, "json_ref_key": row.json_ref_key,
        "raw_relative_path": row.raw_relative_path if row.paths_configured else "",
        "labeled_relative_path": row.labeled_relative_path if row.paths_configured else "",
        "thumbnail_enabled": settings.thumbnail_enabled,
        "thumbnail_error": settings.thumbnail_error,
        "annotation_method_code": row.annotation_method_code,
    }


@app.put("/api/settings")
def update_settings(payload: SettingsPayload, db: Session = Depends(get_db)) -> dict:
    try:
        resolve_directory(settings.data_root_path, payload.raw_relative_path)
        resolve_directory(settings.data_root_path, payload.labeled_relative_path)
        if payload.thumbnail_relative_path and len(Path(payload.thumbnail_relative_path).parts) != 1:
            raise ValueError("썸네일 폴더는 데이터 루트 바로 아래의 최상위 폴더를 선택하세요.")
        if payload.thumbnail_relative_path and settings.thumbnail_enabled:
            resolve_directory(settings.thumbnail_root_path, payload.thumbnail_relative_path)
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    if payload.raw_relative_path == payload.labeled_relative_path:
        raise HTTPException(status_code=422, detail="원천데이터와 라벨데이터 폴더는 서로 달라야 합니다.")
    row = get_or_create_settings(db)
    roots_changed = (
        row.raw_relative_path != payload.raw_relative_path
        or row.labeled_relative_path != payload.labeled_relative_path
    )
    if roots_changed:
        stale_candidates = db.scalars(
            select(SelectionCandidate).where(SelectionCandidate.selection_status != SelectionStatus.SELECTED)
        ).all()
        for candidate in stale_candidates:
            db.delete(candidate)
    row.mapping_strategy = payload.mapping_strategy
    row.json_ref_key = payload.json_ref_key
    row.raw_relative_path = payload.raw_relative_path
    row.labeled_relative_path = payload.labeled_relative_path
    row.thumbnail_relative_path = payload.thumbnail_relative_path
    row.annotation_method_code = payload.annotation_method_code
    row.paths_configured = True
    db.commit()
    return retrieve_settings(db)


@app.get("/api/directories")
def list_directories(path: str = Query(default=""), root: str = Query(default="data")) -> dict:
    if root not in {"data", "thumbnail"}:
        raise HTTPException(status_code=400, detail="지원하지 않는 폴더 루트입니다.")
    if root == "thumbnail" and not settings.thumbnail_enabled:
        raise HTTPException(status_code=409, detail=settings.thumbnail_error)
    browse_root = settings.thumbnail_root_path if root == "thumbnail" else settings.data_root_path
    relative = path.strip().replace("\\", "/").strip("/")
    try:
        current = browse_root.resolve() if not relative else resolve_directory(browse_root, relative)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    root = browse_root.resolve()
    children = []
    try:
        for child in sorted(current.iterdir(), key=lambda item: item.name.lower()):
            if child.name.startswith(".") or child.is_symlink():
                continue
            if child.is_dir():
                children.append({"name": child.name, "path": child.relative_to(root).as_posix()})
            elif child.is_file():
                continue
    except OSError as exc:
        raise HTTPException(status_code=409, detail=f"폴더 목록을 읽을 수 없습니다: {exc}") from exc
    current_relative = "" if current == root else current.relative_to(root).as_posix()
    parent = "" if not current_relative else Path(current_relative).parent.as_posix()
    return {
        "root_container_path": root.as_posix(),
        "root_host_path": settings.host_thumbnail_root_path if root == settings.thumbnail_root_path.resolve() else settings.host_data_root_path,
        "current": current_relative,
        "parent": "" if parent == "." else parent,
        "directories": children,
        "file_counts": {},
        "root_type": root,
    }


def _count_directory_files(path: Path) -> dict[str, int]:
    counts: dict[str, int] = {}
    for child in path.iterdir():
        if child.name.startswith(".") or child.is_symlink() or not child.is_file():
            continue
        extension = child.suffix.lower() or "(확장자 없음)"
        counts[extension] = counts.get(extension, 0) + 1
    return dict(sorted(counts.items()))


@app.get("/api/directories/file-counts")
async def directory_file_counts(path: str = Query(default=""), root: str = Query(default="data")) -> dict:
    if root not in {"data", "thumbnail"}:
        raise HTTPException(status_code=400, detail="지원하지 않는 폴더 루트입니다.")
    if root == "thumbnail" and not settings.thumbnail_enabled:
        raise HTTPException(status_code=409, detail=settings.thumbnail_error)
    browse_root = settings.thumbnail_root_path if root == "thumbnail" else settings.data_root_path
    relative = path.strip().replace("\\", "/").strip("/")
    try:
        current = browse_root.resolve() if not relative else resolve_directory(browse_root, relative)
        counts = await asyncio.to_thread(_count_directory_files, current)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except OSError as exc:
        raise HTTPException(status_code=409, detail=f"파일 목록을 읽을 수 없습니다: {exc}") from exc
    return {"path": relative, "file_counts": counts}


def _volume_status(key: str, label: str, path: Path, host_path: str) -> dict:
    resolved = path.resolve(strict=False)
    exists = resolved.exists()
    total = used = free = 0
    if exists:
        try:
            usage = shutil.disk_usage(resolved)
            total, used, free = usage.total, usage.used, usage.free
        except OSError:
            pass
    return {
        "key": key,
        "label": label,
        "host_path": host_path,
        "container_path": resolved.as_posix(),
        "exists": exists,
        "is_mount": resolved.is_mount() if exists else False,
        "readable": exists and os.access(resolved, os.R_OK),
        "writable": exists and os.access(resolved, os.W_OK),
        "total_bytes": total,
        "used_bytes": used,
        "free_bytes": free,
    }


@app.get("/api/volumes")
def volume_status(db: Session = Depends(get_db)) -> dict:
    app_setting = get_or_create_settings(db)
    volumes = [
        _volume_status("data_root", "데이터 루트", settings.data_root_path, settings.host_data_root_path),
        *([_volume_status("thumbnail_root", "썸네일 루트", settings.thumbnail_root_path, settings.host_thumbnail_root_path)] if settings.thumbnail_root_path else []),
        _volume_status("selected", "선별 결과", settings.selected_data_path, settings.host_selected_data_path),
        _volume_status("app_data", "애플리케이션 데이터", settings.app_data_path, settings.host_app_data_path),
    ]
    selected_directories = []
    for key, label, relative in (
        ("raw", "원천데이터 폴더", app_setting.raw_relative_path if app_setting.paths_configured else ""),
        ("labeled", "라벨데이터 폴더", app_setting.labeled_relative_path if app_setting.paths_configured else ""),
    ):
        try:
            if not relative:
                raise ValueError("폴더가 선택되지 않았습니다.")
            path = resolve_directory(settings.data_root_path, relative)
            ready, readable, writable = True, os.access(path, os.R_OK), os.access(path, os.W_OK)
        except ValueError:
            ready, readable, writable = False, False, False
        selected_directories.append({
            "key": key, "label": label, "relative_path": relative,
            "exists": ready, "readable": readable, "writable": writable,
        })
    return {"volumes": volumes, "selected_directories": selected_directories}


@app.post("/api/scan")
def scan(payload: ScanRequest, db: Session = Depends(get_db)) -> dict:
    app_setting = get_or_create_settings(db)
    strategy = payload.mapping_strategy or app_setting.mapping_strategy
    json_ref_key = (payload.json_ref_key or app_setting.json_ref_key).strip()
    annotation_method_code = payload.annotation_method_code or app_setting.annotation_method_code
    if payload.raw_relative_path:
        app_setting.raw_relative_path = payload.raw_relative_path.strip().replace("\\", "/").strip("/")
    if payload.labeled_relative_path:
        app_setting.labeled_relative_path = payload.labeled_relative_path.strip().replace("\\", "/").strip("/")
    if payload.thumbnail_relative_path is not None:
        app_setting.thumbnail_relative_path = payload.thumbnail_relative_path.strip().replace("\\", "/").strip("/")
    if payload.raw_relative_path and payload.labeled_relative_path:
        app_setting.paths_configured = True
    if strategy == MappingStrategy.JSON_REF_KEY and not json_ref_key:
        raise HTTPException(status_code=422, detail="JSON 참조키를 입력하세요.")
    app_setting.mapping_strategy = strategy
    app_setting.json_ref_key = json_ref_key
    app_setting.annotation_method_code = annotation_method_code
    try:
        raw_root, labeled_root = configured_data_roots(settings, app_setting)
    except ValueError as exc:
        db.rollback()
        raise HTTPException(status_code=422, detail=str(exc)) from exc

    # 호스트 바인드 마운트 파일시스템의 일시적 잠금/readonly 오류에 대비해 재시도한다.
    retry_delays = (0.3, 0.8)
    last_error: SQLAlchemyError | None = None
    for delay in (*retry_delays, None):
        try:
            result = scan_candidates(db, settings, strategy, json_ref_key, raw_root, labeled_root)
            if settings.thumbnail_enabled:
                app_setting = get_or_create_settings(db)
                result["thumbnail_queued"] = queue_thumbnail_generation(
                    settings, "", raw_root,
                    [item.relative for item in discover_files(raw_root)],
                )
            else:
                result["thumbnail_queued"] = 0
            return result
        except SQLAlchemyError as exc:
            db.rollback()
            last_error = exc
            if delay is None:
                break
            time.sleep(delay)
    raise HTTPException(status_code=409, detail=f"후보 중복 또는 DB 갱신 오류: {last_error}") from last_error


@app.get("/api/stats")
def stats(db: Session = Depends(get_db)) -> dict:
    match_rows = db.execute(
        select(SelectionCandidate.match_status, func.count()).group_by(SelectionCandidate.match_status)
    ).all()
    match_summary = {status.value: 0 for status in MatchStatus}
    match_summary.update({str(status): int(count) for status, count in match_rows})
    return {"selection": selection_summary(db), "match": match_summary}


@app.get("/api/candidates")
def list_candidates(
    selection_status: str | None = Query(default=None),
    match_status: str | None = Query(default=None),
    search: str | None = Query(default=None),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=20, ge=1, le=100),
    db: Session = Depends(get_db),
) -> dict:
    app_setting = get_or_create_settings(db)
    raw_root: Path | None = None
    if app_setting.paths_configured:
        try:
            raw_root, _ = configured_data_roots(settings, app_setting)
        except ValueError:
            raw_root = None
    query = select(SelectionCandidate)
    if selection_status:
        query = query.where(SelectionCandidate.selection_status == selection_status)
    if match_status:
        query = query.where(SelectionCandidate.match_status == match_status)
    if search and search.strip():
        pattern = f"%{search.strip()}%"
        query = query.where(or_(
            SelectionCandidate.match_key.ilike(pattern),
            SelectionCandidate.files.any(SelectionCandidateFile.original_relative_path.ilike(pattern)),
        ))

    count = int(db.scalar(select(func.count()).select_from(query.subquery())) or 0)
    rows = db.scalars(
        query.options(selectinload(SelectionCandidate.files))
        .order_by(SelectionCandidate.created_at.desc(), SelectionCandidate.id)
        .offset((page - 1) * page_size)
        .limit(page_size)
    ).all()
    return {
        "results": [_serialize_candidate(row, raw_root=raw_root, thumbnail_relative_path="") for row in rows],
        "count": count,
        "page": page,
        "page_size": page_size,
        "total_pages": math.ceil(count / page_size) if count else 0,
        "summary": selection_summary(db),
    }


@app.get("/api/candidates/{candidate_id}")
def candidate_detail(candidate_id: str, db: Session = Depends(get_db)) -> dict:
    candidate = _candidate_or_404(db, candidate_id)
    if candidate.selection_status == SelectionStatus.SELECTED:
        labeled_root = settings.data_root_path
    else:
        _, labeled_root = configured_data_roots(settings, get_or_create_settings(db))
    raw_root, _ = configured_data_roots(settings, get_or_create_settings(db))
    app_setting = get_or_create_settings(db)
    return _serialize_candidate(candidate, include_label=True, labeled_root=labeled_root, raw_root=raw_root, thumbnail_relative_path="")


@app.post("/api/candidates/{candidate_id}/decision")
def decide_candidate(candidate_id: str, payload: DecisionRequest, db: Session = Depends(get_db)) -> dict:
    candidate = _candidate_or_404(db, candidate_id)
    if candidate.selection_status == SelectionStatus.SELECTED:
        if payload.decision == SelectionStatus.SELECTED:
            return _serialize_candidate(candidate, include_label=True, labeled_root=settings.data_root_path)
        raise HTTPException(status_code=409, detail="복사 완료된 후보는 제외 상태로 변경할 수 없습니다.")
    raw_root, labeled_root = configured_data_roots(settings, get_or_create_settings(db))

    if payload.decision == SelectionStatus.REJECTED:
        candidate.selection_status = SelectionStatus.REJECTED
        candidate.decided_at = utcnow()
        candidate.error_message = ""
        db.commit()
        return _serialize_candidate(candidate, include_label=True, labeled_root=labeled_root)

    copied = []
    try:
        copied = copy_candidate_files(candidate, settings, raw_root, labeled_root, overwrite=payload.overwrite)
        candidate.selection_status = SelectionStatus.SELECTED
        candidate.decided_at = utcnow()
        candidate.error_message = ""
        db.commit()
        finalize_copies(copied)
    except CopyConflictError as exc:
        db.rollback()
        raise HTTPException(status_code=409, detail={
            "code": "copy_conflict",
            "message": "같은 경로의 파일이 이미 있습니다. 덮어쓸지 확인하세요.",
            "conflicts": exc.conflicts,
        }) from exc
    except Exception as exc:
        db.rollback()
        if copied:
            rollback_copies(copied)
        candidate = _candidate_or_404(db, candidate_id)
        candidate.selection_status = SelectionStatus.MOVE_FAILED
        candidate.error_message = str(exc)
        db.commit()
        raise HTTPException(status_code=409, detail=f"파일 복사 실패: {exc}") from exc
    return _serialize_candidate(candidate, include_label=True, labeled_root=labeled_root)


@app.post("/api/candidates/{candidate_id}/reset")
def reset_candidate(candidate_id: str, db: Session = Depends(get_db)) -> dict:
    candidate = _candidate_or_404(db, candidate_id)
    if candidate.selection_status == SelectionStatus.SELECTED:
        raise HTTPException(status_code=409, detail="복사 완료된 후보는 초기화할 수 없습니다.")
    candidate.selection_status = SelectionStatus.PENDING
    candidate.decided_at = None
    candidate.error_message = "" if candidate.match_status == MatchStatus.MATCHED else candidate.error_message
    db.commit()
    _, labeled_root = configured_data_roots(settings, get_or_create_settings(db))
    return _serialize_candidate(candidate, include_label=True, labeled_root=labeled_root)


@app.get("/api/files/{group}")
def stream_file(group: str, path: str = Query(min_length=1), db: Session = Depends(get_db)):
    if group == "selected":
        root = settings.selected_data_path
    elif group in {"raw", "labeled"}:
        raw_root, labeled_root = configured_data_roots(settings, get_or_create_settings(db))
        root = raw_root if group == "raw" else labeled_root
    else:
        raise HTTPException(status_code=404, detail="지원하지 않는 파일 그룹입니다.")
    try:
        target = resolve_safe(root, path)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except FileNotFoundError as exc:
        raise HTTPException(status_code=404, detail="파일을 찾을 수 없습니다.") from exc
    return FileResponse(target, filename=target.name)


@app.get("/api/thumbnails")
def stream_thumbnail(path: str = Query(min_length=1), db: Session = Depends(get_db)):
    app_setting = get_or_create_settings(db)
    if not settings.thumbnail_enabled:
        raise HTTPException(status_code=409, detail=settings.thumbnail_error)
    try:
        raw_root, _ = configured_data_roots(settings, app_setting)
        target = find_thumbnail(settings, "", raw_root, path)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    if target is None:
        raise HTTPException(status_code=404, detail="썸네일을 찾을 수 없습니다.")
    return FileResponse(target, filename=target.name)


@app.post("/api/thumbnails/generate/{candidate_id}")
def generate_candidate_thumbnails(candidate_id: str, db: Session = Depends(get_db)):
    candidate = _candidate_or_404(db, candidate_id)
    app_setting = get_or_create_settings(db)
    logger.info("thumbnail generation requested: candidate_id=%s enabled=%s root=%s", candidate_id, settings.thumbnail_enabled, settings.thumbnail_root_path)
    if not settings.thumbnail_enabled:
        logger.error("thumbnail generation disabled: candidate_id=%s reason=%s", candidate_id, settings.thumbnail_error)
        raise HTTPException(status_code=409, detail=settings.thumbnail_error)
    raw_root, _ = configured_data_roots(settings, app_setting)
    files = [item.original_relative_path for item in candidate.files if item.file_group == "raw"]
    logger.debug("thumbnail generation candidates: candidate_id=%s raw_root=%s files=%s", candidate_id, raw_root, files)
    try:
        queued = queue_thumbnail_generation(settings, "", raw_root, files)
    except Exception:
        logger.exception("thumbnail generation queue failed: candidate_id=%s raw_root=%s files=%s", candidate_id, raw_root, files)
        raise
    logger.info("thumbnail generation queued: candidate_id=%s total=%d queued=%d", candidate_id, len(files), queued)
    return {"queued": queued}


@app.get("/api/thumbnails/progress")
def thumbnail_progress(db: Session = Depends(get_db)) -> dict:
    if not settings.thumbnail_enabled:
        return {"active": False, "total": 0, "completed": 0, "generating": 0, "percent": 0}
    app_setting = get_or_create_settings(db)
    try:
        raw_root, _ = configured_data_roots(settings, app_setting)
    except ValueError:
        return {"active": False, "total": 0, "completed": 0, "generating": 0, "percent": 0}
    total = completed = generating = pending = queued_pending = 0
    for entry in discover_files(raw_root):
        if not entry.is_image:
            continue
        total += 1
        state, _ = thumbnail_state(settings, "", raw_root, entry.relative, entry.absolute)
        job_state = thumbnail_job_status(entry.relative)
        if state == "available":
            completed += 1
        elif job_state == "generating" or (job_state is None and state == "generating"):
            generating += 1
        else:
            pending += 1
            if job_state == "pending":
                queued_pending += 1
    return {
        "active": generating > 0 or queued_pending > 0,
        "total": total,
        "completed": completed,
        "generating": generating,
        "pending": pending,
        "percent": round(completed / total * 100) if total else 100,
    }


@app.get("/api/candidates/{candidate_id}/files/{file_id}/text")
def preview_text_file(candidate_id: str, file_id: int, db: Session = Depends(get_db)) -> dict:
    candidate = _candidate_or_404(db, candidate_id)
    item = next((file for file in candidate.files if file.id == file_id), None)
    if item is None:
        raise HTTPException(status_code=404, detail="후보에 포함된 파일을 찾을 수 없습니다.")

    if item.selected_relative_path:
        root, relative = settings.selected_data_path, item.selected_relative_path
    else:
        raw_root, labeled_root = configured_data_roots(settings, get_or_create_settings(db))
        root = raw_root if item.file_group == "raw" else labeled_root
        relative = item.original_relative_path
    try:
        target = resolve_safe(root, relative)
        preview_limit = 512 * 1024
        with target.open("rb") as stream:
            content = stream.read(preview_limit + 1)
    except (ValueError, FileNotFoundError) as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except OSError as exc:
        raise HTTPException(status_code=409, detail=f"파일을 읽을 수 없습니다: {exc}") from exc

    truncated = len(content) > preview_limit
    content = content[:preview_limit]
    if b"\x00" in content:
        return {"previewable": False, "reason": "바이너리 데이터가 포함되어 텍스트로 표시할 수 없습니다."}

    decoded = None
    encoding = ""
    for candidate_encoding in ("utf-8-sig", "cp949"):
        try:
            decoded = content.decode(candidate_encoding)
            encoding = "utf-8" if candidate_encoding == "utf-8-sig" else candidate_encoding
            break
        except UnicodeDecodeError:
            if truncated:
                decoded = content.decode(candidate_encoding, errors="ignore")
                encoding = "utf-8" if candidate_encoding == "utf-8-sig" else candidate_encoding
                break
            continue
    if decoded is None:
        return {"previewable": False, "reason": "지원하는 문자 인코딩(UTF-8, CP949)으로 해석할 수 없습니다."}
    printable = sum(character.isprintable() or character in "\r\n\t" for character in decoded)
    if decoded and printable / len(decoded) < 0.85:
        return {"previewable": False, "reason": "제어 문자가 많아 텍스트 파일로 판단할 수 없습니다."}
    return {
        "previewable": True,
        "content": decoded,
        "encoding": encoding,
        "truncated": truncated,
        "preview_bytes": len(content),
        "size": target.stat().st_size,
    }


def _format_datetime(value: datetime | None) -> str:
    return value.astimezone().strftime("%Y-%m-%d %H:%M:%S") if value and value.tzinfo else (value.strftime("%Y-%m-%d %H:%M:%S") if value else "")


def _style_sheet(sheet) -> None:
    fill = PatternFill("solid", fgColor="1D4ED8")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in sheet.columns:
        letter = column[0].column_letter
        width = min(max(len(str(cell.value or "")) for cell in column) + 2, 70)
        sheet.column_dimensions[letter].width = width


@app.get("/api/selections/export.xlsx")
def export_selections(db: Session = Depends(get_db)):
    rows = db.scalars(
        candidate_query()
        .where(SelectionCandidate.selection_status == SelectionStatus.SELECTED)
        .order_by(SelectionCandidate.decided_at.desc())
    ).all()

    workbook = Workbook()
    selection_sheet = workbook.active
    selection_sheet.title = "선별 결과"
    selection_sheet.append([
        "Selection ID", "매칭 키", "매칭 방식", "상태", "원천 파일 수",
        "원천 상대 경로 목록", "라벨 상대 경로", "결정 일시",
    ])
    for candidate in rows:
        raw_paths = [item.original_relative_path for item in candidate.files if item.file_group == "raw"]
        label_paths = [item.original_relative_path for item in candidate.files if item.file_group == "labeled"]
        selection_sheet.append([
            candidate.id, candidate.match_key, candidate.mapping_strategy, candidate.selection_status,
            len(raw_paths), "\n".join(raw_paths), "\n".join(label_paths), _format_datetime(candidate.decided_at),
        ])
        selection_sheet.cell(selection_sheet.max_row, 6).alignment = Alignment(wrap_text=True, vertical="top")
        selection_sheet.cell(selection_sheet.max_row, 7).alignment = Alignment(wrap_text=True, vertical="top")
    _style_sheet(selection_sheet)

    file_sheet = workbook.create_sheet("파일 목록")
    file_sheet.append([
        "Selection ID", "파일 그룹", "참조 순서", "원본 상대 경로", "복사 후 상대 경로",
        "확장자", "크기(bytes)", "수정 시간",
    ])
    for candidate in rows:
        for item in candidate.files:
            file_sheet.append([
                candidate.id, item.file_group, item.reference_order, item.original_relative_path,
                item.selected_relative_path, item.extension, item.size,
                datetime.fromtimestamp(item.mtime).strftime("%Y-%m-%d %H:%M:%S"),
            ])
    _style_sheet(file_sheet)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"prework_selections_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
