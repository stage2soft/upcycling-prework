from __future__ import annotations

from io import BytesIO
import json

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import delete

from app.database import Base, SessionLocal, engine
from app.main import app, settings
from app.models import SelectionCandidate


def test_candidate_pagination_selection_and_excel_export():
    Base.metadata.create_all(engine)
    with SessionLocal() as db:
        db.execute(delete(SelectionCandidate))
        db.commit()

    raw = settings.data_root_path / "raw/region/images/api-scene.png"
    obj = settings.data_root_path / "raw/region/objects/api-scene.pcd"
    label = settings.data_root_path / "labeled/region/labels/api-scene.json"
    raw.parent.mkdir(parents=True, exist_ok=True)
    obj.parent.mkdir(parents=True, exist_ok=True)
    label.parent.mkdir(parents=True, exist_ok=True)
    raw.write_bytes(b"png-placeholder")
    obj.write_bytes(b"pcd-placeholder")
    label.write_text(json.dumps({"objects": [{"bbox": [1, 2, 3, 4]}]}), encoding="utf-8")

    with TestClient(app) as client:
        directories = client.get("/api/directories")
        assert directories.status_code == 200
        assert directories.json()["root_container_path"] == settings.data_root_path.resolve().as_posix()
        assert directories.json()["root_host_path"] == settings.host_data_root_path
        assert {item["name"] for item in directories.json()["directories"]} >= {"raw", "labeled"}
        assert client.get("/api/directories", params={"path": "../"}).status_code == 400

        saved = client.put("/api/settings", json={
            "mapping_strategy": "file_name", "json_ref_key": "data_key",
            "raw_relative_path": "raw", "labeled_relative_path": "labeled",
        })
        assert saved.status_code == 200
        assert saved.json()["raw_relative_path"] == "raw"

        volumes = client.get("/api/volumes")
        assert volumes.status_code == 200
        assert {item["key"] for item in volumes.json()["volumes"]} == {"data_root", "selected", "app_data"}
        assert all(item["host_path"] for item in volumes.json()["volumes"])
        assert all(item["exists"] for item in volumes.json()["selected_directories"])

        scan = client.post("/api/scan", json={"mapping_strategy": "file_name", "json_ref_key": "data_key"})
        assert scan.status_code == 200
        listing = client.get("/api/candidates", params={"page": 1, "page_size": 1, "match_status": "matched"})
        assert listing.status_code == 200
        payload = listing.json()
        assert payload["count"] == 1
        assert payload["page_size"] == 1
        candidate_id = payload["results"][0]["id"]

        detail = client.get(f"/api/candidates/{candidate_id}")
        assert detail.status_code == 200
        assert len([item for item in detail.json()["files"] if item["file_group"] == "raw"]) == 2
        object_file = next(item for item in detail.json()["files"] if item["extension"] == ".pcd")
        text_preview = client.get(f"/api/candidates/{candidate_id}/files/{object_file['id']}/text")
        assert text_preview.status_code == 200
        assert text_preview.json()["previewable"] is True
        assert text_preview.json()["content"] == "pcd-placeholder"

        selected_raw = settings.selected_data_path / "raw/region/images/api-scene.png"
        selected_raw.parent.mkdir(parents=True, exist_ok=True)
        selected_raw.write_bytes(b"existing-file")

        conflict = client.post(f"/api/candidates/{candidate_id}/decision", json={"decision": "selected"})
        assert conflict.status_code == 409
        assert conflict.json()["detail"]["code"] == "copy_conflict"
        assert conflict.json()["detail"]["conflicts"] == ["raw/region/images/api-scene.png"]
        assert selected_raw.read_bytes() == b"existing-file"

        decision = client.post(
            f"/api/candidates/{candidate_id}/decision",
            json={"decision": "selected", "overwrite": True},
        )
        assert decision.status_code == 200
        assert decision.json()["selection_status"] == "selected"
        assert raw.is_file()
        assert obj.is_file()
        assert label.is_file()
        assert selected_raw.read_bytes() == b"png-placeholder"
        assert (settings.selected_data_path / "raw/region/objects/api-scene.pcd").is_file()
        assert (settings.selected_data_path / "labeled/region/labels/api-scene.json").is_file()
        assert not list(selected_raw.parent.glob(".api-scene.png.prework-*.bak"))

        export = client.get("/api/selections/export.xlsx")
        assert export.status_code == 200
        workbook = load_workbook(BytesIO(export.content), read_only=True)
        assert workbook.sheetnames == ["선별 결과", "파일 목록"]
        assert workbook["선별 결과"].max_row == 2
        assert workbook["파일 목록"].max_row == 4
